from flask import request


class Result():
    result_according_template = ""
    results_list = []

    def __init__(self):
        result_according_template = ""
        results_list = []

    def generate_results(self, creative_rules, macros, creative_path, export_path):
        import shutil, os
        from pathlib import Path
        ''' example
        creative_rules = [
            { #第一组选择十个创意的规则
                "name": "##cpn##_HK_Signup_##size##",
                "creative_files": ["160x600.png", "300x50.png", "300x250.png", "300x600.png", "320x50.png",
                                   "320x100.png", "320x480.png", "728x90.png", "1024x768.png", "1200x628.png"],
                "clickthrough_url": "https://www.bybit.com/zh-TW/register?regEmail&medium=paid_displayp&source=TTD&channel=paid_&campaign=HK_##cpn##&term=All_##size##&content=V1&dtpid=1634637492338",
                "landing_page_url": "https://www.bybit.com/"
            },
            {  #第二组选择十个创意的规则
                "creative_files": ["1 (1).png", "1 (2).png", "1 (3).png"],
                name：  Control_20230213_##creativeFileName##  （第一组）
                Description:  Ecoflow 2023 Control
                Asset File Name:  Control_20230213_##creativeFileName##   (不输入则不需要重命名，保留图片后缀)
                ClickUL:  https://us.ecoflow.com/?utm_source=tzjs&utm_medium=display&utm_campaign=controlled
                landing：  https://us.ecoflow.com/
            }
            ,      
            {
            "creative_files": request.form.get('files').split(","),
            "name": request.form.get('creative_name_rule'),
            "Description": request.form.get('Description_rule'),
            "Asset_File_Name": request.form.get('Asset_File_Name_rule'),
            "clickthrough_url": request.form.get('clickthrough_rule'),
            "landing_page_url": request.form.get('landing_page_rule'),
            "Adgroup_name": request.form.get('Ag_list').split(",")
        }
        ]
        '''

        results = []
        for rule_dict in creative_rules:
            sheet = rule_dict["creative_type"]
            # hosted native
            Native_Short_Title = rule_dict['Native_Short_Title']
            Native_Long_Title = rule_dict['Native_Long_Title']
            Native_Short_Description = rule_dict['Native_Short_Description']
            Native_Long_Description = rule_dict['Native_Long_Description']
            Sponsor = rule_dict['Sponsor']
            # extra for display+video
            Impression_Tracking_url = rule_dict['Impression_Tracking_url']
            # 3 rd Video Audio
            VAST_XML_URL = rule_dict['VAST_XML_URL']
            # hosted flash+html5
            Click_Tracking_Parameter = rule_dict['Click_Tracking_Parameter']
            Width = rule_dict['Width']
            Height = rule_dict['Height']
            # 3rd display
            AdTag = rule_dict['AdTag']
            if len(rule_dict["Adgroup_name"]) == 0:
                for creative in rule_dict["creative_files"]:
                    creative_file_name = "{}".format(Path(creative).stem)
                    creative_file_ext = os.path.splitext(creative)[1]
                    Asset_File_Name = rule_dict["Asset_File_Name"].replace(
                        "{}".format(macros["creative_file_name"]["key"]),
                        creative_file_name)
                    # Asset_File_Name.replace( "{}".format( macros["creative_file_name"]["key"] ), creative_file_name )
                    # Asset_File_Name = rule_dict["Asset_File_Name"]
                    # if Asset_File_Name != creative_file_name:
                    #     shutil.copy(os.path.join(creative_path, creative), os.path.join(os.path.abspath(export_path),
                    #                                                                     "{}{}".format(Asset_File_Name,
                    #                                                                                   creative_file_ext)))
                    #     '''
                    #     if os.path.exists( os.path.join( os.path.abspath( export_path ), Asset_File_Name)  ):
                    #         os.remove( os.path.abspath( "{}{}".format( export_path, Asset_File_Name) ))
                    #     os.rename( os.path.join(os.path.abspath( export_path ), creative), os.path.join( os.path.abspath( export_path ), Asset_File_Name))
                    #     '''
                    #
                    #     # flask.send_file(os.path.join(os.path.abspath( "./export" ), Asset_File_Name))
                    # else:
                    #     shutil.copy(os.path.join(creative_path, creative),
                    #                 os.path.join(os.path.abspath(export_path), creative))
                    lines = {}
                    name = rule_dict["name"]
                    clickthrough_url = rule_dict["clickthrough_url"]
                    landing_page_url = rule_dict["landing_page_url"]
                    Description = rule_dict["Description"]
                    # Asset_File_Name = "{}{}".format(Asset_File_Name, creative_file_ext)
                    for macro_obj, macro_dict in macros.items():
                        if macro_dict['dynamic'] == False:
                            name = name.replace("{}".format(macro_dict["key"]), macro_dict["value"])
                            clickthrough_url = clickthrough_url.replace("{}".format(macro_dict["key"]),
                                                                        macro_dict["value"])
                            landing_page_url = landing_page_url.replace("{}".format(macro_dict["key"]),
                                                                        macro_dict["value"])
                            Description = Description.replace("{}".format(macro_dict["key"]), macro_dict["value"])
                            Asset_File_Name = Asset_File_Name.replace("{}".format(macro_dict["key"]),
                                                                      macro_dict["value"])
                        else:  # 动态宏替换处理
                            if macro_obj == "Creative_size":
                                # 读创意文件的size, 存如 size
                                # image
                                from PIL import Image
                                import filetype
                                path = "{}\\{}".format(creative_path, creative)
                                kind = filetype.guess(path)
                                if kind.mime.split('/')[0] == 'image':
                                    im = Image.open(path)
                                    print("{}\\{}".format(creative_path, creative))
                                    print(creative)
                                    size = "{}x{}".format(im.size[0], im.size[1])
                                    # size = '160x40'  # example
                                # video
                                import cv2
                                if kind.mime.split('/')[0] == 'video':
                                    vid = cv2.VideoCapture(path)
                                    h = vid.get(cv2.CAP_PROP_FRAME_HEIGHT)
                                    w = vid.get(cv2.CAP_PROP_FRAME_WIDTH)
                                    size = "{}x{}".format(h, w)
                                # other file type
                                else:
                                    size = 'N/A'
                                name = name.replace("{}".format(macro_dict["key"]), size)
                                clickthrough_url = clickthrough_url.replace("{}".format(macro_dict["key"]), size)
                                landing_page_url = landing_page_url.replace("{}".format(macro_dict["key"]), size)
                                Description = Description.replace("{}".format(macro_dict["key"]), size)
                                Asset_File_Name = Asset_File_Name.replace("{}".format(macro_dict["key"]), size)
                            if macro_obj == "Creative_width":
                                # read width
                                # image
                                from PIL import Image
                                import filetype
                                path = "{}\\{}".format(creative_path, creative)
                                kind = filetype.guess(path)
                                if kind.mime.split('/')[0] == 'image':
                                    im = Image.open(path)
                                    print("{}\\{}".format(creative_path, creative))
                                    print(creative)
                                    width = "{}".format(im.size[0])
                                    # size = '160x40' width='160' # example
                                # video
                                import cv2
                                if kind.mime.split('/')[0] == 'video':
                                    vid = cv2.VideoCapture(path)
                                    width = vid.get(cv2.CAP_PROP_FRAME_WIDTH)
                                    width = "{}".format(width)

                                # other file type
                                else:
                                    width = 'N/A'
                                name = name.replace("{}".format(macro_dict["key"]), width)
                                clickthrough_url = clickthrough_url.replace("{}".format(macro_dict["key"]), width)
                                landing_page_url = landing_page_url.replace("{}".format(macro_dict["key"]), width)
                                Description = Description.replace("{}".format(macro_dict["key"]), width)
                                Asset_File_Name = Asset_File_Name.replace("{}".format(macro_dict["key"]), width)
                                Width = Width.replace("{}".format(macro_dict["key"]), width)

                            if macro_obj == "Creative_height":
                                # read height
                                # image
                                from PIL import Image
                                import filetype
                                path = "{}\\{}".format(creative_path, creative)
                                kind = filetype.guess(path)
                                if kind.mime.split('/')[0] == 'image':
                                    im = Image.open(path)
                                    print("{}\\{}".format(creative_path, creative))
                                    print(creative)
                                    height = "{}".format(im.size[1])
                                    # size = '160x40' width='40' # example
                                # video
                                import cv2
                                if kind.mime.split('/')[0] == 'video':
                                    vid = cv2.VideoCapture(path)
                                    height = vid.get(cv2.CAP_PROP_FRAME_HEIGHT)
                                    height = "{}".format(height)

                                # other file type
                                else:
                                    height = 'N/A'
                                name = name.replace("{}".format(macro_dict["key"]), height)
                                clickthrough_url = clickthrough_url.replace("{}".format(macro_dict["key"]), height)
                                landing_page_url = landing_page_url.replace("{}".format(macro_dict["key"]), height)
                                Description = Description.replace("{}".format(macro_dict["key"]), height)
                                Asset_File_Name = Asset_File_Name.replace("{}".format(macro_dict["key"]), height)
                                Height = Height.replace("{}".format(macro_dict["key"]),height)

                            if macro_obj == "creative_file_name":
                                # 读文件名 存入creative_file_name
                                name = name.replace("{}".format(macro_dict["key"]), Asset_File_Name)
                                clickthrough_url = clickthrough_url.replace("{}".format(macro_dict["key"]),
                                                                            Asset_File_Name)
                                landing_page_url = landing_page_url.replace("{}".format(macro_dict["key"]),
                                                                            Asset_File_Name)
                                Description = Description.replace("{}".format(macro_dict["key"]), Asset_File_Name)
                                Asset_File_Name = Asset_File_Name.replace("{}".format(macro_dict["key"]),
                                                                          Asset_File_Name)
                            # if macro_obj == "Adgroup":
                            #     AdgroupList = request.POST.get('Adgroup_list')
                            #     # AdgroupList = []
                            #     name = name.replace("{}".format(macro_dict["key"]), AdgroupList)
                            #     clickthrough_url = clickthrough_url.replace("{}".format(macro_dict["key"]), AdgroupList)
                            #     landing_page_url = landing_page_url.replace("{}".format(macro_dict["key"]), AdgroupList)
                            #     Description = Description.replace("{}".format(macro_dict["key"]), AdgroupList)
                            #     Asset_File_Name = Asset_File_Name.replace("{}".format(macro_dict["key"]), AdgroupList)
                    Asset_File_Name = "{}{}".format(Asset_File_Name, creative_file_ext)
                    if Asset_File_Name != creative_file_name:
                        shutil.copy(os.path.join(creative_path, creative), os.path.join(os.path.abspath(export_path),
                                                                                        Asset_File_Name))
                        '''
                        if os.path.exists( os.path.join( os.path.abspath( export_path ), Asset_File_Name)  ):
                            os.remove( os.path.abspath( "{}{}".format( export_path, Asset_File_Name) ))
                        os.rename( os.path.join(os.path.abspath( export_path ), creative), os.path.join( os.path.abspath( export_path ), Asset_File_Name))
                        '''

                        # flask.send_file(os.path.join(os.path.abspath( "./export" ), Asset_File_Name))
                    else:
                        shutil.copy(os.path.join(creative_path, creative),
                                    os.path.join(os.path.abspath(export_path), creative))
                    lines['sheet'] = sheet  # get sheet name by creative type
                    lines['Name'] = name  # name
                    lines['Description'] = Description  # Description
                    lines['Asset_File_Name'] = Asset_File_Name  # Asset File Name (required)
                    lines['Clickthrough_URL'] = clickthrough_url  # Clickthrough URL (required)
                    lines['Landing_Page_URL'] = landing_page_url  # Landing Page URL (required)
                    # hosted native
                    lines['Native_Short_Title'] = Native_Short_Title
                    lines['Native_Long_Title'] = Native_Long_Title
                    lines['Native_Short_Description'] = Native_Short_Description
                    lines['Native_Long_Description'] = Native_Long_Description
                    lines['Sponsor'] = Sponsor
                    # extra
                    lines['Impression_Tracking_url'] = Impression_Tracking_url
                    # 3 rd Video Audio
                    lines['VAST_XML_URL'] = VAST_XML_URL
                    # hosted flash+html5
                    lines['Click_Tracking_Parameter'] = Click_Tracking_Parameter
                    lines['Width'] = Width
                    lines['Height'] = Height
                    # 3rd display
                    lines['AdTag'] = AdTag
                    results.append(lines)
            else:
                for adgroup in rule_dict["Adgroup_name"]:
                    for creative in rule_dict["creative_files"]:
                        creative_file_name = "{}".format(Path(creative).stem)
                        creative_file_ext = os.path.splitext(creative)[1]
                        Asset_File_Name = rule_dict["Asset_File_Name"].replace(
                            "{}".format(macros["creative_file_name"]["key"]),
                            creative_file_name)
                        lines = {}
                        name = rule_dict["name"]
                        clickthrough_url = rule_dict["clickthrough_url"]
                        landing_page_url = rule_dict["landing_page_url"]
                        Description = rule_dict["Description"]
                        # Asset_File_Name = "{}{}".format(Asset_File_Name, creative_file_ext)
                        for macro_obj, macro_dict in macros.items():
                            if macro_dict['dynamic'] == False:
                                name = name.replace("{}".format(macro_dict["key"]), macro_dict["value"])
                                clickthrough_url = clickthrough_url.replace("{}".format(macro_dict["key"]),
                                                                            macro_dict["value"])
                                landing_page_url = landing_page_url.replace("{}".format(macro_dict["key"]),
                                                                            macro_dict["value"])
                                Description = Description.replace("{}".format(macro_dict["key"]), macro_dict["value"])
                                Asset_File_Name = Asset_File_Name.replace("{}".format(macro_dict["key"]),
                                                                          macro_dict["value"])
                            else:  # 动态宏替换处理
                                if macro_obj == "Creative_size":
                                    # 读创意文件的size, 存如 size
                                    # image
                                    from PIL import Image
                                    import filetype
                                    path = "{}\\{}".format(creative_path, creative)
                                    kind = filetype.guess(path)
                                    if kind.mime.split('/')[0] == 'image':
                                        im = Image.open(path)
                                        print("{}\\{}".format(creative_path, creative))
                                        print(creative)
                                        size = "{}x{}".format(im.size[0], im.size[1])
                                        # size = '160x40'  # example
                                    # video
                                    import cv2
                                    if kind.mime.split('/')[0] == 'video':
                                        vid = cv2.VideoCapture(path)
                                        height = vid.get(cv2.CAP_PROP_FRAME_HEIGHT)
                                        width = vid.get(cv2.CAP_PROP_FRAME_WIDTH)
                                        size = "{}x{}".format(height, width)
                                    # other file type
                                    else:
                                        size = 'N/A'
                                    name = name.replace("{}".format(macro_dict["key"]), size)
                                    clickthrough_url = clickthrough_url.replace("{}".format(macro_dict["key"]), size)
                                    landing_page_url = landing_page_url.replace("{}".format(macro_dict["key"]), size)
                                    Description = Description.replace("{}".format(macro_dict["key"]), size)
                                    Asset_File_Name = Asset_File_Name.replace("{}".format(macro_dict["key"]), size)
                                if macro_obj == "Creative_width":
                                    # read width
                                    # image
                                    from PIL import Image
                                    import filetype
                                    path = "{}\\{}".format(creative_path, creative)
                                    kind = filetype.guess(path)
                                    if kind.mime.split('/')[0] == 'image':
                                        im = Image.open(path)
                                        print("{}\\{}".format(creative_path, creative))
                                        print(creative)
                                        width = "{}".format(im.size[0])
                                        # size = '160x40' width='160' # example
                                    # video
                                    import cv2
                                    if kind.mime.split('/')[0] == 'video':
                                        vid = cv2.VideoCapture(path)
                                        width = vid.get(cv2.CAP_PROP_FRAME_WIDTH)
                                        width = "{}".format(width)

                                    # other file type
                                    else:
                                        width = 'N/A'
                                    name = name.replace("{}".format(macro_dict["key"]), width)
                                    clickthrough_url = clickthrough_url.replace("{}".format(macro_dict["key"]), width)
                                    landing_page_url = landing_page_url.replace("{}".format(macro_dict["key"]), width)
                                    Description = Description.replace("{}".format(macro_dict["key"]), width)
                                    Asset_File_Name = Asset_File_Name.replace("{}".format(macro_dict["key"]), width)
                                    Width = Width.replace("{}".format(macro_dict["key"]), width)

                                if macro_obj == "Creative_height":
                                    # read height
                                    # image
                                    from PIL import Image
                                    import filetype
                                    path = "{}\\{}".format(creative_path, creative)
                                    kind = filetype.guess(path)
                                    if kind.mime.split('/')[0] == 'image':
                                        im = Image.open(path)
                                        print("{}\\{}".format(creative_path, creative))
                                        print(creative)
                                        height = "{}".format(im.size[1])
                                        # size = '160x40' width='40' # example
                                    # video
                                    import cv2
                                    if kind.mime.split('/')[0] == 'video':
                                        vid = cv2.VideoCapture(path)
                                        height = vid.get(cv2.CAP_PROP_FRAME_HEIGHT)
                                        height = "{}".format(height)

                                    # other file type
                                    else:
                                        height = 'N/A'
                                    name = name.replace("{}".format(macro_dict["key"]), height)
                                    clickthrough_url = clickthrough_url.replace("{}".format(macro_dict["key"]), height)
                                    landing_page_url = landing_page_url.replace("{}".format(macro_dict["key"]), height)
                                    Description = Description.replace("{}".format(macro_dict["key"]), height)
                                    Asset_File_Name = Asset_File_Name.replace("{}".format(macro_dict["key"]), height)
                                    Height = Height.replace("{}".format(macro_dict["key"]), height)

                                if macro_obj == "creative_file_name":
                                    # 读文件名 存入creative_file_name
                                    name = name.replace("{}".format(macro_dict["key"]), Asset_File_Name)
                                    clickthrough_url = clickthrough_url.replace("{}".format(macro_dict["key"]),
                                                                                Asset_File_Name)
                                    landing_page_url = landing_page_url.replace("{}".format(macro_dict["key"]),
                                                                                Asset_File_Name)
                                    Description = Description.replace("{}".format(macro_dict["key"]), Asset_File_Name)
                                    Asset_File_Name = Asset_File_Name.replace("{}".format(macro_dict["key"]),
                                                                              Asset_File_Name)
                                if macro_obj == "Adgroup":
                                    # AdgroupList = []
                                    name = name.replace("{}".format(macro_dict["key"]), adgroup)
                                    clickthrough_url = clickthrough_url.replace("{}".format(macro_dict["key"]), adgroup)
                                    landing_page_url = landing_page_url.replace("{}".format(macro_dict["key"]), adgroup)
                                    Description = Description.replace("{}".format(macro_dict["key"]), adgroup)
                                    Asset_File_Name = Asset_File_Name.replace("{}".format(macro_dict["key"]), adgroup)
                        Asset_File_Name = "{}{}".format(Asset_File_Name, creative_file_ext)
                        if Asset_File_Name != creative_file_name:
                            shutil.copy(os.path.join(creative_path, creative),
                                        os.path.join(os.path.abspath(export_path),
                                                     Asset_File_Name))
                            '''
                            if os.path.exists( os.path.join( os.path.abspath( export_path ), Asset_File_Name)  ):
                                os.remove( os.path.abspath( "{}{}".format( export_path, Asset_File_Name) ))
                            os.rename( os.path.join(os.path.abspath( export_path ), creative), os.path.join( os.path.abspath( export_path ), Asset_File_Name))
                            '''

                            # flask.send_file(os.path.join(os.path.abspath( "./export" ), Asset_File_Name))
                        else:
                            shutil.copy(os.path.join(creative_path, creative),
                                        os.path.join(os.path.abspath(export_path), creative))

                        lines['sheet'] = sheet  # get sheet name by creative type
                        lines['Name'] = name  # name
                        lines['Description'] = Description  # Description
                        lines['Asset_File_Name'] = Asset_File_Name  # Asset File Name (required)
                        lines['Clickthrough_URL'] = clickthrough_url  # Clickthrough URL (required)
                        lines['Landing_Page_URL'] = landing_page_url  # Landing Page URL (required)
                        # hosted native
                        lines['Native_Short_Title'] = Native_Short_Title
                        lines['Native_Long_Title'] = Native_Long_Title
                        lines['Native_Short_Description'] = Native_Short_Description
                        lines['Native_Long_Description'] = Native_Long_Description
                        lines['Sponsor'] = Sponsor
                        # extra
                        lines['Impression_Tracking_url'] = Impression_Tracking_url
                        # 3 rd Video Audio
                        lines['VAST_XML_URL'] = VAST_XML_URL
                        # hosted flash+html5
                        lines['Click_Tracking_Parameter'] = Click_Tracking_Parameter
                        lines['Width'] = Width
                        lines['Height'] = Height
                        # 3rd display
                        lines['AdTag'] = AdTag
                        results.append(lines)

        self.results_list = results
        return True

    def save_display_sheet(self, workbook, data_dict, row, path, name, fill_bgcolor_and_str=0, fill_cell="",
                           color="#fff", fill_str=""):
        sheet = workbook.get_sheet_by_name('Hosted Display')
        sheet_field_index = {
            "Name": 0,
            "Description": 1,
            "Asset_File_Name": 2,
            "Clickthrough_URL": 3,
            "Landing_Page_URL": 4,
            # "Third_Party_Tracking_URL_1": 10
            "Impression_Tracking_url": 10
        }
        for k, v in data_dict.items():
            if k != 'sheet':
                try:
                    sheet.cell(row=row + 1, column=sheet_field_index[k] + 1, value=v)
                except Exception:
                    print(k, v, row, path, name)

        if fill_bgcolor_and_str == 1:
            self.cell_fill(sheet=sheet, row=row + 1, column=sheet_field_index[fill_cell] + 1, color=color,
                           fill_str=fill_str)
        new_row = row + 1
        return new_row

    def save_flash_sheet(self, workbook, data_dict, row, path, name, fill_bgcolor_and_str=0, fill_cell="", color="#fff",
                         fill_str=""):
        sheet = workbook.get_sheet_by_name('Hosted Flash')
        sheet_field_index = {
            "Name": 0,
            "Description": 1,
            "Asset_File_Name": 2,
            "Clickthrough_URL": 4,
            "Landing_Page_URL": 5,
            "Click_Tracking_Parameter": 6

        }
        for k, v in data_dict.items():
            if k != 'sheet':
                try:
                    sheet.cell(row=row + 1, column=sheet_field_index[k] + 1, value=v)
                except Exception:
                    print(k, v, row, path, name)

        if fill_bgcolor_and_str == 1:
            self.cell_fill(sheet=sheet, row=row + 1, column=sheet_field_index[fill_cell] + 1, color=color,
                           fill_str=fill_str)
        new_row = row + 1
        return new_row

    def save_html5_sheet(self, workbook, data_dict, row, path, name, fill_bgcolor_and_str=0, fill_cell="", color="#fff",
                         fill_str=""):
        sheet = workbook.get_sheet_by_name('Hosted HTML 5')
        sheet_field_index = {
            "Name": 0,
            "Description": 1,
            "Width": 2,
            "Height": 3,
            "Asset_File_Name": 4,
            "Clickthrough_URL": 6,
            "Landing_Page_URL": 8,
            "Click_Tracking_Parameter": 9

        }
        for k, v in data_dict.items():
            if k != 'sheet':
                try:
                    sheet.cell(row=row + 1, column=sheet_field_index[k] + 1, value=v)
                except Exception:
                    print(k, v, row, path, name)

        if fill_bgcolor_and_str == 1:
            self.cell_fill(sheet=sheet, row=row + 1, column=sheet_field_index[fill_cell] + 1, color=color,
                           fill_str=fill_str)
        new_row = row + 1
        return new_row

    def save_native_sheet(self, workbook, data_dict, row, path, name, fill_bgcolor_and_str=0, fill_cell="",
                          color="#fff", fill_str=""):
        sheet = workbook.get_sheet_by_name('Hosted Native')
        sheet_field_index = {
            "Name": 0,
            "Description": 1,
            "Asset_File_Name": 2,
            "Clickthrough_URL": 6,
            "Landing_Page_URL": 7,
            "Native_Short_Title": 10,
            "Native_Long_Title": 11,
            "Native_Short_Description": 12,
            "Native_Long_Description": 13,
            "Sponsor": 14
            # "Third_Party_Tracking_URL_1": 21
            # "impression_url": 21
        }

        for k, v in data_dict.items():
            if k != 'sheet':
                try:
                    sheet.cell(row=row + 1, column=sheet_field_index[k] + 1, value=v)
                except Exception:
                    print(k, v, row, path, name)
        if fill_bgcolor_and_str == 1:
            self.cell_fill(sheet=sheet, row=row + 1, column=sheet_field_index[fill_cell] + 1, color=color,
                           fill_str=fill_str)
        new_row = row + 1
        return new_row

    # by Tony
    def save_video_sheet(self, workbook, data_dict, row, path, name, fill_bgcolor_and_str=0, fill_cell="", color="#fff",
                         fill_str=""):
        sheet = workbook.get_sheet_by_name('Hosted Video')
        sheet_field_index = {
            "Name": 0,
            "Description": 1,
            "Asset_File_Name": 2,
            "Clickthrough_URL": 3,
            "Landing_Page_URL": 4,
            # "Video_Event_URLs_Impression": 9
            "Impression_Tracking_url": 9
            # "Video_Event_URLs_Click": 10,
            # "Video_Event_URLs_Complete": 15
        }

        for k, v in data_dict.items():
            if k != 'sheet':
                try:
                    sheet.cell(row=row + 1, column=sheet_field_index[k] + 1, value=v)
                except Exception:
                    print(k, v, row, path, name)
        if fill_bgcolor_and_str == 1:
            self.cell_fill(sheet=sheet, row=row + 1, column=sheet_field_index[fill_cell] + 1, color=color,
                           fill_str=fill_str)
        new_row = row + 1
        return new_row

    def save_audio_sheet(self, workbook, data_dict, row, path, name, fill_bgcolor_and_str=0, fill_cell="", color="#fff",
                         fill_str=""):
        sheet = workbook.get_sheet_by_name('Hosted Audio')
        sheet_field_index = {
            "Name": 0,
            "Description": 1,
            "Asset_File_Name": 2,
            "Clickthrough_URL": 3,
            "Landing_Page_URL": 4

        }

        for k, v in data_dict.items():
            if k != 'sheet':
                try:
                    sheet.cell(row=row + 1, column=sheet_field_index[k] + 1, value=v)
                except Exception:
                    print(k, v, row, path, name)
        if fill_bgcolor_and_str == 1:
            self.cell_fill(sheet=sheet, row=row + 1, column=sheet_field_index[fill_cell] + 1, color=color,
                           fill_str=fill_str)
        new_row = row + 1
        return new_row

    def save_3p_display_sheet(self, workbook, data_dict, row, path, name, fill_bgcolor_and_str=0, fill_cell="",
                              color="#fff", fill_str=""):
        sheet = workbook.get_sheet_by_name('Third Party Display')
        sheet_field_index = {
            "Name": 0,
            "Description": 1,
            "Width": 2,
            "Height": 3,
            "AdTag": 4,
            "Landing_Page_URL": 5
        }
        for k, v in data_dict.items():
            if k != 'sheet':
                try:
                    sheet.cell(row=row + 1, column=sheet_field_index[k] + 1, value=v)
                except Exception:
                    print(k, v, row, path, name)

        if fill_bgcolor_and_str == 1:
            self.cell_fill(sheet=sheet, row=row + 1, column=sheet_field_index[fill_cell] + 1, color=color,
                           fill_str=fill_str)
        new_row = row + 1
        return new_row

    def save_3p_video_sheet(self, workbook, data_dict, row, path, name, fill_bgcolor_and_str=0, fill_cell="",
                            color="#fff", fill_str=""):
        sheet = workbook.get_sheet_by_name('Third Party Video')
        sheet_field_index = {
            "Name": 0,
            "Description": 1,
            "VAST_XML_URL": 2,
            "Landing_Page_URL": 3
        }

        for k, v in data_dict.items():
            if k != 'sheet':
                try:
                    sheet.cell(row=row + 1, column=sheet_field_index[k] + 1, value=v)
                except Exception:
                    print(k, v, row, path, name)
        if fill_bgcolor_and_str == 1:
            self.cell_fill(sheet=sheet, row=row + 1, column=sheet_field_index[fill_cell] + 1, color=color,
                           fill_str=fill_str)
        new_row = row + 1
        return new_row

    def save_3p_audio_sheet(self, workbook, data_dict, row, path, name, fill_bgcolor_and_str=0, fill_cell="",
                            color="#fff", fill_str=""):
        sheet = workbook.get_sheet_by_name('Third Party Audio')
        sheet_field_index = {
            "Name": 0,
            "Description": 1,
            "VAST_XML_URL": 2,
            "Landing_Page_URL": 3

        }

        for k, v in data_dict.items():
            if k != 'sheet':
                try:
                    sheet.cell(row=row + 1, column=sheet_field_index[k] + 1, value=v)
                except Exception:
                    print(k, v, row, path, name)
        if fill_bgcolor_and_str == 1:
            self.cell_fill(sheet=sheet, row=row + 1, column=sheet_field_index[fill_cell] + 1, color=color,
                           fill_str=fill_str)
        new_row = row + 1
        return new_row

    def cell_fill(self, sheet, row, column, color, fill_str=""):
        from openpyxl.styles import PatternFill
        orange_fill = PatternFill(fill_type='solid', fgColor=color)
        sheet.cell(row=row, column=column, value="_".join(fill_str))
        sheet.cell(row=row, column=column).fill = orange_fill

    # def export_to_excel(self, path, name, sheet="Hosted Display", data_set={}):
    def export_to_excel(self, path, name, data_set={}):
        import openpyxl
        # 写入正式数据
        display_row = 1
        video_row = 1
        native_row = 1
        flash_row = 1
        html_row = 1
        audio_row = 1
        display3_row = 1
        video3_row = 1
        audio3_row = 1
        workbook = openpyxl.load_workbook('./static/BulkCreativeImportTemplate.v28.xlsx')

        if len(data_set) == 0:
            data_set = self.results_list
        for dicts in data_set:
            if dicts['sheet'] == "Hosted Display":
                display_row = self.save_display_sheet(workbook, dicts, display_row, path, name)
            elif dicts['sheet'] == "Hosted Flash":
                flash_row = self.save_flash_sheet(workbook, dicts, flash_row, path, name)
            elif dicts['sheet'] == "Hosted Native":
                native_row = self.save_native_sheet(workbook, dicts, native_row, path, name)
            elif dicts['sheet'] == "Hosted HTML5":
                html_row = self.save_html5_sheet(workbook, dicts, html_row, path, name)
            elif dicts['sheet'] == "Hosted Video":
                video_row = self.save_video_sheet(workbook, dicts, video_row, path, name)
            elif dicts['sheet'] == "Hosted Audio":
                audio_row = self.save_audio_sheet(workbook, dicts, audio_row, path, name)
            elif dicts['sheet'] == "Third Party Display":
                display3_row = self.save_3p_display_sheet(workbook, dicts, display3_row, path, name)
            elif dicts['sheet'] == "Third Party Video":
                video3_row = self.save_3p_video_sheet(workbook, dicts, video3_row, path, name)
            elif dicts['sheet'] == "Third Party Audio":
                audio3_row = self.save_3p_audio_sheet(workbook, dicts, audio3_row, path, name)
            else:
                print(dicts)
                pass
        workbook.save("{}/{}".format(path, name))
        return name


class MiaozhenResult(Result):

    def save_results_from_miaozhen_set(self, outside_data_set, export_path, result_file, name_conversion_from={},
                                       name_sep="-"):
        import openpyxl, re
        kv_sheets = {
            "Mobile Video": ["Hosted Video", 1],
            "Display": ["Hosted Display", 1],
            "Native Display": ["Hosted Native", 1],
            "Splash": ["Hosted Display", 1]
        }
        # from urllib.parse import urlparse
        workbook = openpyxl.load_workbook('./BulkCreativeImportTemplate.v28.xlsx')

        results = []

        # 根据 name_conversion_from，设置单元格底色
        # 创意名称单元个底色初始化
        Asset_File_Name_colors = []
        current_name_content = []  # 根绝name_conversion_from 数量，设置对应元素，如果 name_conversion_from 有四项，那么本变量中就有四个元素
        cell_colors = ["D9E1F2", "FCE4D6"]
        cell_color_index = 0
        for line_index, data_lines in enumerate(outside_data_set):
            lines = {}
            name = data_lines[2]
            Description = ""
            Asset_File_Name = ""  # need modified

            try:
                clickthrough_url = data_lines[4]
            except IndexError:
                print("Line:{} doesn't have click_url".format(line_index, data_lines))

            impression_url = data_lines[3]

            landing_Page_url = clickthrough_url.split("&o=")[1]

            lines['Name'] = name  # name
            lines['Description'] = Description  # Description
            lines['Asset_File_Name'] = Asset_File_Name  # Asset File Name (required)
            lines['Clickthrough_URL'] = clickthrough_url  # Clickthrough URL (required)
            lines['Landing_Page_URL'] = landing_Page_url  # Landing Page URL (required)
            lines['impression_url'] = impression_url  # Third Party Tracking URL 1
            results.append(lines)

        # 根据关键词区分sheet
        all_keywords = []
        for field, data_set in name_conversion_from.items():
            all_keywords += data_set
        re_pattern_string = "|".join(all_keywords)

        # 写入正式数据
        display_row = 1
        video_row = 1
        native_row = 1
        name_keywords = []
        previous_name_keywords = []
        for l_index, lines in enumerate(results):

            # 从lines[0](name)中找关键字匹配sheet
            ad_type = re.findall(r'{}({}){}'.format(name_sep, "|".join(kv_sheets.keys()), name_sep), lines['Name'])
            if len(ad_type) > 0:
                sheet = kv_sheets[ad_type[0]][0]  # 要处理的sheet名称
            name_keywords = re.findall(r'{}'.format(re_pattern_string), lines['Name'])
            if name_keywords != previous_name_keywords:
                cell_color_index += 1
                cell_color = cell_colors[cell_color_index % 2]

            if sheet == "Hosted Display":
                display_row = self.save_display_sheet(workbook, lines, display_row, export_path, result_file,
                                                      fill_bgcolor_and_str=1, fill_cell="Asset_File_Name",
                                                      color=cell_color, fill_str=name_keywords)
            elif sheet == "Hosted Video":
                video_row = self.save_video_sheet(workbook, lines, video_row, export_path, result_file,
                                                  fill_bgcolor_and_str=1, fill_cell="Asset_File_Name", color=cell_color,
                                                  fill_str=name_keywords)
            elif sheet == "Hosted Native":
                native_row = self.save_native_sheet(workbook, lines, native_row, export_path, result_file,
                                                    fill_bgcolor_and_str=1, fill_cell="Asset_File_Name",
                                                    color=cell_color, fill_str=name_keywords)
            else:
                print(l_index, lines, results)
                pass
            previous_name_keywords = name_keywords
        workbook.save("{}/{}".format(export_path, result_file))

        return result_file
